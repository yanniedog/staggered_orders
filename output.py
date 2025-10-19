"""
Unified output module consolidating visualization and Excel export functionality.
Combines visualizer.py and excel_export.py into a single comprehensive output module.
"""
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import pandas as pd
import numpy as np
from typing import Dict, Tuple, List
import os
import seaborn as sns
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from logger import log_info, log_warning, log_error, log_debug


def create_output_dir():
    """Create output directory if it doesn't exist"""
    try:
        if not os.path.exists('output'):
            os.makedirs('output')
    except Exception as e:
        log_warning(f"Warning: Could not create output directory: {e}")


# ============================================================================
# VISUALIZATION FUNCTIONS
# ============================================================================

def plot_probability_fit_quality_dashboard(depths: np.ndarray, empirical_probs: np.ndarray,
                                         theta: float, p: float, fit_metrics: Dict,
                                         depths_upward: np.ndarray, empirical_probs_upward: np.ndarray,
                                         theta_sell: float, p_sell: float, fit_metrics_sell: Dict) -> None:
    """
    Create comprehensive probability fit quality dashboard (2x2 grid).
    
    Shows buy-side and sell-side Weibull fits with residuals, Q-Q plots, and confidence intervals.
    """
    try:
        create_output_dir()
        
        # Validate inputs
        if len(depths) == 0 or len(empirical_probs) == 0:
            log_warning("Warning: Empty data arrays provided to probability fit dashboard")
            return
        
        if np.any(np.isnan(depths)) or np.any(np.isnan(empirical_probs)):
            log_warning("Warning: NaN values in probability fit data")
            return
        
        if np.any(np.isnan(depths_upward)) or np.any(np.isnan(empirical_probs_upward)):
            log_warning("Warning: NaN values in upward probability fit data")
            return
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
        
        # Plot 1: Buy-side Weibull fit with residuals
        fitted_probs = np.exp(-(depths / theta) ** p)
        residuals = empirical_probs - fitted_probs
        
        ax1.plot(depths, empirical_probs, 'bo', alpha=0.6, label='Empirical', markersize=4)
        ax1.plot(depths, fitted_probs, 'r-', linewidth=2, label='Weibull Fit')
        
        # Add confidence intervals if available
        if 'prediction_intervals' in fit_metrics:
            pred_intervals = fit_metrics['prediction_intervals']
            lower_bounds = [interval[0] for interval in pred_intervals]
            upper_bounds = [interval[1] for interval in pred_intervals]
            ax1.fill_between(depths, lower_bounds, upper_bounds, alpha=0.2, color='red', label='95% CI')
        
        ax1.set_xlabel('Depth (%)')
        ax1.set_ylabel('Touch Probability')
        ax1.set_title(f'Buy-Side Weibull Fit\nR² = {fit_metrics["r_squared"]:.3f}, Quality: {fit_metrics.get("fit_quality", "unknown")}')
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        
        # Plot 2: Sell-side Weibull fit with residuals
        fitted_probs_sell = np.exp(-(depths_upward / theta_sell) ** p_sell)
        residuals_sell = empirical_probs_upward - fitted_probs_sell
        
        ax2.plot(depths_upward, empirical_probs_upward, 'go', alpha=0.6, label='Empirical', markersize=4)
        ax2.plot(depths_upward, fitted_probs_sell, 'orange', linewidth=2, label='Weibull Fit')
        
        # Add confidence intervals if available
        if 'prediction_intervals' in fit_metrics_sell:
            pred_intervals_sell = fit_metrics_sell['prediction_intervals']
            lower_bounds_sell = [interval[0] for interval in pred_intervals_sell]
            upper_bounds_sell = [interval[1] for interval in pred_intervals_sell]
            ax2.fill_between(depths_upward, lower_bounds_sell, upper_bounds_sell, alpha=0.2, color='orange', label='95% CI')
        
        ax2.set_xlabel('Depth (%)')
        ax2.set_ylabel('Touch Probability')
        ax2.set_title(f'Sell-Side Weibull Fit\nR² = {fit_metrics_sell["r_squared"]:.3f}, Quality: {fit_metrics_sell.get("fit_quality", "unknown")}')
        ax2.legend()
        ax2.grid(True, alpha=0.3)
        
        # Plot 3: Q-Q plot for buy-side residuals
        stats.probplot(residuals, dist="norm", plot=ax3)
        ax3.set_title('Buy-Side Residuals Q-Q Plot')
        ax3.grid(True, alpha=0.3)
        
        # Plot 4: Q-Q plot for sell-side residuals
        stats.probplot(residuals_sell, dist="norm", plot=ax4)
        ax4.set_title('Sell-Side Residuals Q-Q Plot')
        ax4.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig('output/probability_fit_quality_dashboard.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        log_info("Probability fit quality dashboard saved to output/probability_fit_quality_dashboard.png")
        
    except Exception as e:
        log_error(f"Error creating probability fit quality dashboard: {e}")
        plt.close('all')


def plot_ladder_configuration_summary(depths: np.ndarray, allocations: np.ndarray,
                                    paired_orders_df: pd.DataFrame, current_price: float) -> None:
    """
    Create ladder configuration summary visualization.
    
    Shows depth distribution, allocation distribution, and profit potential.
    """
    try:
        create_output_dir()
        
        if len(depths) == 0 or len(allocations) == 0:
            log_warning("Warning: Empty data arrays provided to ladder configuration summary")
            return
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
        
        # Plot 1: Depth distribution
        ax1.bar(range(len(depths)), depths, color='skyblue', edgecolor='black', alpha=0.7)
        ax1.set_xlabel('Rung Number')
        ax1.set_ylabel('Depth (%)')
        ax1.set_title('Ladder Depth Distribution')
        ax1.grid(True, alpha=0.3)
        
        # Plot 2: Allocation distribution
        ax2.bar(range(len(allocations)), allocations, color='lightcoral', edgecolor='black', alpha=0.7)
        ax2.set_xlabel('Rung Number')
        ax2.set_ylabel('Allocation ($)')
        ax2.set_title('Size Allocation Distribution')
        ax2.grid(True, alpha=0.3)
        
        # Plot 3: Profit potential
        if not paired_orders_df.empty and 'profit_pct' in paired_orders_df.columns:
            ax3.bar(range(len(paired_orders_df)), paired_orders_df['profit_pct'], 
                   color='lightgreen', edgecolor='black', alpha=0.7)
            ax3.set_xlabel('Rung Number')
            ax3.set_ylabel('Profit (%)')
            ax3.set_title('Profit Potential per Rung')
            ax3.grid(True, alpha=0.3)
        
        # Plot 4: Price levels
        buy_prices = current_price * (1 - depths / 100)
        if not paired_orders_df.empty and 'sell_price' in paired_orders_df.columns:
            sell_prices = paired_orders_df['sell_price']
            ax4.plot(range(len(buy_prices)), buy_prices, 'bo-', label='Buy Prices', linewidth=2, markersize=6)
            ax4.plot(range(len(sell_prices)), sell_prices, 'ro-', label='Sell Prices', linewidth=2, markersize=6)
            ax4.axhline(y=current_price, color='black', linestyle='--', alpha=0.7, label='Current Price')
            ax4.set_xlabel('Rung Number')
            ax4.set_ylabel('Price ($)')
            ax4.set_title('Buy/Sell Price Levels')
            ax4.legend()
            ax4.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig('output/ladder_configuration_summary.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        log_info("Ladder configuration summary saved to output/ladder_configuration_summary.png")
        
    except Exception as e:
        log_error(f"Error creating ladder configuration summary: {e}")
        plt.close('all')


def plot_expected_performance_metrics(scenarios_df: pd.DataFrame, optimal_scenario: Dict) -> None:
    """
    Create expected performance metrics visualization.
    
    Shows key performance indicators and scenario comparison.
    """
    try:
        create_output_dir()
        
        if scenarios_df is None or scenarios_df.empty:
            log_warning("Warning: Empty scenarios DataFrame provided to performance metrics")
            return
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
        
        # Plot 1: Expected monthly profit vs profit target
        if 'profit_target_pct' in scenarios_df.columns and 'expected_monthly_profit' in scenarios_df.columns:
            ax1.scatter(scenarios_df['profit_target_pct'], scenarios_df['expected_monthly_profit'], 
                       alpha=0.7, s=50, c='blue', edgecolors='black')
            ax1.set_xlabel('Profit Target (%)')
            ax1.set_ylabel('Expected Monthly Profit ($)')
            ax1.set_title('Expected Monthly Profit vs Profit Target')
            ax1.grid(True, alpha=0.3)
        
        # Plot 2: Expected timeframe vs profit target
        if 'profit_target_pct' in scenarios_df.columns and 'expected_timeframe_hours' in scenarios_df.columns:
            ax2.scatter(scenarios_df['profit_target_pct'], scenarios_df['expected_timeframe_hours'], 
                       alpha=0.7, s=50, c='red', edgecolors='black')
            ax2.set_xlabel('Profit Target (%)')
            ax2.set_ylabel('Expected Timeframe (hours)')
            ax2.set_title('Expected Timeframe vs Profit Target')
            ax2.grid(True, alpha=0.3)
        
        # Plot 3: Capital efficiency distribution
        if 'expected_profit_per_dollar_per_month' in scenarios_df.columns:
            ax3.hist(scenarios_df['expected_profit_per_dollar_per_month'], bins=20, 
                    color='green', alpha=0.7, edgecolor='black')
            ax3.set_xlabel('Expected Monthly Return per Dollar')
            ax3.set_ylabel('Frequency')
            ax3.set_title('Capital Efficiency Distribution')
            ax3.grid(True, alpha=0.3)
        
        # Plot 4: Risk-return scatter
        if 'sharpe_ratio' in scenarios_df.columns and 'expected_monthly_profit' in scenarios_df.columns:
            ax4.scatter(scenarios_df['sharpe_ratio'], scenarios_df['expected_monthly_profit'], 
                       alpha=0.7, s=50, c='purple', edgecolors='black')
            ax4.set_xlabel('Sharpe Ratio')
            ax4.set_ylabel('Expected Monthly Profit ($)')
            ax4.set_title('Risk-Return Profile')
            ax4.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig('output/expected_performance_metrics.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        log_info("Expected performance metrics saved to output/expected_performance_metrics.png")
        
    except Exception as e:
        log_error(f"Error creating expected performance metrics: {e}")
        plt.close('all')


def plot_profitability_distribution(paired_orders_df: pd.DataFrame) -> None:
    """
    Create profitability distribution visualization.
    
    Shows profit distribution across rungs and overall statistics.
    """
    try:
        create_output_dir()
        
        if paired_orders_df.empty or 'profit_pct' not in paired_orders_df.columns:
            log_warning("Warning: No profit data available for profitability distribution")
            return
        
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))
        
        # Plot 1: Profit distribution histogram
        ax1.hist(paired_orders_df['profit_pct'], bins=20, color='lightgreen', alpha=0.7, edgecolor='black')
        ax1.set_xlabel('Profit per Pair (%)')
        ax1.set_ylabel('Frequency')
        ax1.set_title('Profit Distribution Across Rungs')
        ax1.grid(True, alpha=0.3)
        
        # Add statistics
        mean_profit = paired_orders_df['profit_pct'].mean()
        median_profit = paired_orders_df['profit_pct'].median()
        ax1.axvline(mean_profit, color='red', linestyle='--', label=f'Mean: {mean_profit:.2f}%')
        ax1.axvline(median_profit, color='blue', linestyle='--', label=f'Median: {median_profit:.2f}%')
        ax1.legend()
        
        # Plot 2: Profit vs rung number
        ax2.plot(range(len(paired_orders_df)), paired_orders_df['profit_pct'], 
                'bo-', linewidth=2, markersize=6)
        ax2.set_xlabel('Rung Number')
        ax2.set_ylabel('Profit per Pair (%)')
        ax2.set_title('Profit by Rung')
        ax2.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig('output/profitability_distribution.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        log_info("Profitability distribution saved to output/profitability_distribution.png")
        
    except Exception as e:
        log_error(f"Error creating profitability distribution: {e}")
        plt.close('all')


def plot_improved_paired_orders(paired_orders_df: pd.DataFrame, current_price: float) -> None:
    """
    Create improved paired orders visualization.
    
    Shows buy/sell price levels, quantities, and profit margins.
    """
    try:
        create_output_dir()
        
        if paired_orders_df.empty:
            log_warning("Warning: Empty paired orders DataFrame provided")
            return
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
        
        # Plot 1: Price levels
        if 'buy_price' in paired_orders_df.columns and 'sell_price' in paired_orders_df.columns:
            ax1.plot(range(len(paired_orders_df)), paired_orders_df['buy_price'], 
                    'bo-', label='Buy Prices', linewidth=2, markersize=6)
            ax1.plot(range(len(paired_orders_df)), paired_orders_df['sell_price'], 
                    'ro-', label='Sell Prices', linewidth=2, markersize=6)
            ax1.axhline(y=current_price, color='black', linestyle='--', alpha=0.7, label='Current Price')
            ax1.set_xlabel('Rung Number')
            ax1.set_ylabel('Price ($)')
            ax1.set_title('Buy/Sell Price Levels')
            ax1.legend()
            ax1.grid(True, alpha=0.3)
        
        # Plot 2: Quantities
        if 'buy_qty' in paired_orders_df.columns and 'sell_qty' in paired_orders_df.columns:
            ax2.plot(range(len(paired_orders_df)), paired_orders_df['buy_qty'], 
                    'go-', label='Buy Quantities', linewidth=2, markersize=6)
            ax2.plot(range(len(paired_orders_df)), paired_orders_df['sell_qty'], 
                    'mo-', label='Sell Quantities', linewidth=2, markersize=6)
            ax2.set_xlabel('Rung Number')
            ax2.set_ylabel('Quantity')
            ax2.set_title('Buy/Sell Quantities')
            ax2.legend()
            ax2.grid(True, alpha=0.3)
        
        # Plot 3: Notionals
        if 'buy_notional' in paired_orders_df.columns and 'sell_notional' in paired_orders_df.columns:
            ax3.plot(range(len(paired_orders_df)), paired_orders_df['buy_notional'], 
                    'co-', label='Buy Notionals', linewidth=2, markersize=6)
            ax3.plot(range(len(paired_orders_df)), paired_orders_df['sell_notional'], 
                    'yo-', label='Sell Notionals', linewidth=2, markersize=6)
            ax3.set_xlabel('Rung Number')
            ax3.set_ylabel('Notional ($)')
            ax3.set_title('Buy/Sell Notionals')
            ax3.legend()
            ax3.grid(True, alpha=0.3)
        
        # Plot 4: Profit margins
        if 'profit_pct' in paired_orders_df.columns:
            ax4.bar(range(len(paired_orders_df)), paired_orders_df['profit_pct'], 
                   color='lightgreen', edgecolor='black', alpha=0.7)
            ax4.set_xlabel('Rung Number')
            ax4.set_ylabel('Profit Margin (%)')
            ax4.set_title('Profit Margins by Rung')
            ax4.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig('output/improved_paired_orders.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        log_info("Improved paired orders visualization saved to output/improved_paired_orders.png")
        
    except Exception as e:
        log_error(f"Error creating improved paired orders visualization: {e}")
        plt.close('all')


def plot_assumption_validation_panel(scenarios_df: pd.DataFrame, optimal_scenario: Dict) -> None:
    """
    Create assumption validation panel.
    
    Shows validation of key assumptions and model quality.
    """
    try:
        create_output_dir()
        
        if scenarios_df is None or scenarios_df.empty:
            log_warning("Warning: Empty scenarios DataFrame provided to assumption validation")
            return
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
        
        # Plot 1: Touch probability validation
        if 'avg_buy_touch_prob' in scenarios_df.columns and 'avg_sell_touch_prob' in scenarios_df.columns:
            ax1.scatter(scenarios_df['avg_buy_touch_prob'], scenarios_df['avg_sell_touch_prob'], 
                       alpha=0.7, s=50, c='blue', edgecolors='black')
            ax1.set_xlabel('Average Buy Touch Probability')
            ax1.set_ylabel('Average Sell Touch Probability')
            ax1.set_title('Touch Probability Validation')
            ax1.grid(True, alpha=0.3)
        
        # Plot 2: Timeframe reasonableness
        if 'expected_timeframe_hours' in scenarios_df.columns:
            ax2.hist(scenarios_df['expected_timeframe_hours'], bins=20, 
                    color='orange', alpha=0.7, edgecolor='black')
            ax2.set_xlabel('Expected Timeframe (hours)')
            ax2.set_ylabel('Frequency')
            ax2.set_title('Timeframe Distribution')
            ax2.grid(True, alpha=0.3)
        
        # Plot 3: Profit target reasonableness
        if 'profit_target_pct' in scenarios_df.columns:
            ax3.hist(scenarios_df['profit_target_pct'], bins=20, 
                    color='green', alpha=0.7, edgecolor='black')
            ax3.set_xlabel('Profit Target (%)')
            ax3.set_ylabel('Frequency')
            ax3.set_title('Profit Target Distribution')
            ax3.grid(True, alpha=0.3)
        
        # Plot 4: Model quality metrics
        if 'sharpe_ratio' in scenarios_df.columns:
            ax4.hist(scenarios_df['sharpe_ratio'], bins=20, 
                    color='purple', alpha=0.7, edgecolor='black')
            ax4.set_xlabel('Sharpe Ratio')
            ax4.set_ylabel('Frequency')
            ax4.set_title('Risk-Adjusted Return Distribution')
            ax4.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig('output/assumption_validation_panel.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        log_info("Assumption validation panel saved to output/assumption_validation_panel.png")
        
    except Exception as e:
        log_error(f"Error creating assumption validation panel: {e}")
        plt.close('all')


def plot_reality_check_dashboard(scenarios_df: pd.DataFrame, optimal_scenario: Dict) -> None:
    """
    Create reality check dashboard.
    
    Shows practical constraints and market reality checks.
    """
    try:
        create_output_dir()
        
        if scenarios_df is None or scenarios_df.empty:
            log_warning("Warning: Empty scenarios DataFrame provided to reality check")
            return
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
        
        # Plot 1: Market liquidity check
        if 'expected_monthly_fills' in scenarios_df.columns:
            ax1.hist(scenarios_df['expected_monthly_fills'], bins=20, 
                    color='lightblue', alpha=0.7, edgecolor='black')
            ax1.set_xlabel('Expected Monthly Fills')
            ax1.set_ylabel('Frequency')
            ax1.set_title('Market Liquidity Check')
            ax1.grid(True, alpha=0.3)
        
        # Plot 2: Capital efficiency check
        if 'expected_profit_per_dollar_per_month' in scenarios_df.columns:
            ax2.hist(scenarios_df['expected_profit_per_dollar_per_month'], bins=20, 
                    color='lightgreen', alpha=0.7, edgecolor='black')
            ax2.set_xlabel('Expected Monthly Return per Dollar')
            ax2.set_ylabel('Frequency')
            ax2.set_title('Capital Efficiency Check')
            ax2.grid(True, alpha=0.3)
        
        # Plot 3: Risk assessment
        if 'profit_volatility' in scenarios_df.columns:
            ax3.hist(scenarios_df['profit_volatility'], bins=20, 
                    color='lightcoral', alpha=0.7, edgecolor='black')
            ax3.set_xlabel('Profit Volatility')
            ax3.set_ylabel('Frequency')
            ax3.set_title('Risk Assessment')
            ax3.grid(True, alpha=0.3)
        
        # Plot 4: Overall performance
        if 'expected_monthly_profit' in scenarios_df.columns:
            ax4.hist(scenarios_df['expected_monthly_profit'], bins=20, 
                    color='lightyellow', alpha=0.7, edgecolor='black')
            ax4.set_xlabel('Expected Monthly Profit ($)')
            ax4.set_ylabel('Frequency')
            ax4.set_title('Overall Performance')
            ax4.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig('output/reality_check_dashboard.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        log_info("Reality check dashboard saved to output/reality_check_dashboard.png")
        
    except Exception as e:
        log_error(f"Error creating reality check dashboard: {e}")
        plt.close('all')


def create_all_visualizations(depths: np.ndarray, empirical_probs: np.ndarray, 
                            theta: float, p: float, fit_metrics: Dict,
                            ladder_depths: np.ndarray, allocations: np.ndarray, 
                            paired_orders_df: pd.DataFrame,
                            depths_upward: np.ndarray, empirical_probs_upward: np.ndarray, 
                            theta_sell: float, p_sell: float, fit_metrics_sell: Dict,
                            sell_depths: np.ndarray, actual_profits: np.ndarray, 
                            scenarios_df: pd.DataFrame, optimal_scenario: Dict) -> None:
    """
    Create all visualizations for the ladder analysis.
    
    Args:
        depths: Buy-side empirical depths
        empirical_probs: Buy-side empirical probabilities
        theta: Buy-side Weibull scale parameter
        p: Buy-side Weibull shape parameter
        fit_metrics: Buy-side fit metrics
        ladder_depths: Optimized ladder depths
        allocations: Size allocations
        paired_orders_df: Paired orders DataFrame
        depths_upward: Sell-side empirical depths
        empirical_probs_upward: Sell-side empirical probabilities
        theta_sell: Sell-side Weibull scale parameter
        p_sell: Sell-side Weibull shape parameter
        fit_metrics_sell: Sell-side fit metrics
        sell_depths: Sell ladder depths
        actual_profits: Actual profit calculations
        scenarios_df: Scenario analysis results
        optimal_scenario: Optimal scenario details
    """
    log_info("Creating all visualizations...")
    
    try:
        # Create individual visualizations
        plot_probability_fit_quality_dashboard(depths, empirical_probs, theta, p, fit_metrics,
                                             depths_upward, empirical_probs_upward, theta_sell, p_sell, fit_metrics_sell)
        
        plot_ladder_configuration_summary(ladder_depths, allocations, paired_orders_df, 
                                        paired_orders_df['buy_price'].iloc[0] if not paired_orders_df.empty else 100.0)
        
        plot_expected_performance_metrics(scenarios_df, optimal_scenario)
        
        plot_profitability_distribution(paired_orders_df)
        
        plot_improved_paired_orders(paired_orders_df, 
                                  paired_orders_df['buy_price'].iloc[0] if not paired_orders_df.empty else 100.0)
        
        plot_assumption_validation_panel(scenarios_df, optimal_scenario)
        
        plot_reality_check_dashboard(scenarios_df, optimal_scenario)
        
        log_info("All visualizations created successfully")
        
    except Exception as e:
        log_error(f"Error creating visualizations: {e}")


# ============================================================================
# EXCEL EXPORT FUNCTIONS
# ============================================================================

def create_excel_workbook(paired_orders_df: pd.DataFrame, depths: np.ndarray, 
                         allocations: np.ndarray, theta: float, p: float,
                         fit_metrics: Dict, budget: float, current_price: float,
                         theta_sell: float, p_sell: float, fit_metrics_sell: Dict,
                         sell_depths: np.ndarray, actual_profits: np.ndarray,
                         scenarios_df: pd.DataFrame = None,
                         rung_sensitivity_df: pd.DataFrame = None,
                         depth_sensitivity_df: pd.DataFrame = None,
                         combined_sensitivity_df: pd.DataFrame = None) -> None:
    """
    Create comprehensive Excel workbook with paired order specs, analysis, and scenario data.
    """
    try:
        create_output_dir()
        
        # Validate inputs
        if paired_orders_df is None or paired_orders_df.empty:
            log_warning("Warning: Empty paired orders DataFrame provided to Excel export")
            return
        
        if np.any(np.isnan(depths)) or np.any(np.isnan(allocations)):
            log_warning("Warning: NaN values in depths or allocations")
            return
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Executive Summary
        ws_executive = wb.create_sheet("Executive Summary")
        create_executive_summary_sheet(ws_executive, paired_orders_df, fit_metrics, fit_metrics_sell, 
                                     scenarios_df, budget, current_price)
        
        # Sheet 2: Paired Order Specifications
        ws_paired = wb.create_sheet("Paired Orders")
        create_paired_orders_sheet(ws_paired, paired_orders_df)
        
        # Sheet 3: Buy-Side Analysis
        ws_buy_params = wb.create_sheet("Buy-Side Analysis")
        create_buy_parameters_sheet(ws_buy_params, theta, p, fit_metrics, budget, current_price, depths)
        
        # Sheet 4: Sell-Side Analysis
        ws_sell_params = wb.create_sheet("Sell-Side Analysis")
        create_sell_parameters_sheet(ws_sell_params, theta_sell, p_sell, fit_metrics_sell, sell_depths, actual_profits)
        
        # Sheet 5: Size Calculations
        ws_calc = wb.create_sheet("Size Calculations")
        create_calculations_sheet(ws_calc, depths, allocations, theta, p)
        
        # Sheet 6: Profit Analysis
        ws_profit = wb.create_sheet("Profit Analysis")
        create_profit_sheet(ws_profit, paired_orders_df)
        
        # Sheet 7: Touch Probability Analysis
        ws_prob = wb.create_sheet("Touch Probability")
        create_probability_sheet(ws_prob, depths, theta, p)
        
        # Scenario analysis sheets
        if scenarios_df is not None:
            # Sheet 8: Scenario Analysis
            ws_scenarios = wb.create_sheet("Scenario Analysis")
            create_scenario_analysis_sheet(ws_scenarios, scenarios_df)
            
            # Sheet 9: Sensitivity Analysis
            ws_sensitivity = wb.create_sheet("Sensitivity Analysis")
            create_sensitivity_analysis_sheet(ws_sensitivity, rung_sensitivity_df, depth_sensitivity_df)
            
            # Sheet 10: Combined Sensitivity
            ws_combined = wb.create_sheet("Combined Sensitivity")
            create_combined_sensitivity_sheet(ws_combined, combined_sensitivity_df)
        
        # Save workbook
        wb.save('output/ladder_report.xlsx')
        log_info("Excel workbook saved to output/ladder_report.xlsx")
        
    except Exception as e:
        log_error(f"Error creating Excel workbook: {e}")


def create_executive_summary_sheet(ws, paired_orders_df: pd.DataFrame, fit_metrics: Dict, 
                                 fit_metrics_sell: Dict, scenarios_df: pd.DataFrame, 
                                 budget: float, current_price: float) -> None:
    """Create executive summary sheet."""
    try:
        # Title
        ws['A1'] = 'STAGGERED ORDER LADDER ANALYSIS - EXECUTIVE SUMMARY'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Key metrics
        row = 3
        ws[f'A{row}'] = 'Key Metrics'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Total Budget:'
        ws[f'B{row}'] = f'${budget:,.2f}'
        row += 1
        
        ws[f'A{row}'] = 'Current Price:'
        ws[f'B{row}'] = f'${current_price:.2f}'
        row += 1
        
        ws[f'A{row}'] = 'Number of Rungs:'
        ws[f'B{row}'] = len(paired_orders_df)
        row += 1
        
        ws[f'A{row}'] = 'Total Expected Profit:'
        if 'profit_usd' in paired_orders_df.columns:
            ws[f'B{row}'] = f'${paired_orders_df["profit_usd"].sum():,.2f}'
        row += 1
        
        # Weibull fit quality
        row += 1
        ws[f'A{row}'] = 'Model Quality'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Buy-Side R²:'
        ws[f'B{row}'] = f'{fit_metrics.get("r_squared", 0):.4f}'
        row += 1
        
        ws[f'A{row}'] = 'Sell-Side R²:'
        ws[f'B{row}'] = f'{fit_metrics_sell.get("r_squared", 0):.4f}'
        row += 1
        
        # Scenario analysis summary
        if scenarios_df is not None and not scenarios_df.empty:
            row += 1
            ws[f'A{row}'] = 'Scenario Analysis'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws[f'A{row}'] = 'Best Expected Monthly Return:'
            if 'expected_profit_per_dollar_per_month' in scenarios_df.columns:
                best_return = scenarios_df['expected_profit_per_dollar_per_month'].max()
                ws[f'B{row}'] = f'{best_return:.4f}'
            row += 1
            
            ws[f'A{row}'] = 'Average Expected Timeframe:'
            if 'expected_timeframe_hours' in scenarios_df.columns:
                avg_timeframe = scenarios_df['expected_timeframe_hours'].mean()
                ws[f'B{row}'] = f'{avg_timeframe:.1f} hours'
            row += 1
        
    except Exception as e:
        log_error(f"Error creating executive summary sheet: {e}")


def create_paired_orders_sheet(ws, paired_orders_df: pd.DataFrame) -> None:
    """Create paired orders sheet."""
    try:
        # Add headers
        headers = ['Rung', 'Buy Depth (%)', 'Buy Price ($)', 'Buy Qty', 'Buy Notional ($)',
                  'Sell Depth (%)', 'Sell Price ($)', 'Sell Qty', 'Sell Notional ($)',
                  'Profit (%)', 'Profit ($)', 'Expected Profit ($)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Add data
        for row_idx, (_, row) in enumerate(paired_orders_df.iterrows(), 2):
            ws.cell(row=row_idx, column=1).value = row.get('rung', row_idx - 1)
            ws.cell(row=row_idx, column=2).value = row.get('buy_depth_pct', 0)
            ws.cell(row=row_idx, column=3).value = row.get('buy_price', 0)
            ws.cell(row=row_idx, column=4).value = row.get('buy_qty', 0)
            ws.cell(row=row_idx, column=5).value = row.get('buy_notional', 0)
            ws.cell(row=row_idx, column=6).value = row.get('sell_depth_pct', 0)
            ws.cell(row=row_idx, column=7).value = row.get('sell_price', 0)
            ws.cell(row=row_idx, column=8).value = row.get('sell_qty', 0)
            ws.cell(row=row_idx, column=9).value = row.get('sell_notional', 0)
            ws.cell(row=row_idx, column=10).value = row.get('profit_pct', 0)
            ws.cell(row=row_idx, column=11).value = row.get('profit_usd', 0)
            ws.cell(row=row_idx, column=12).value = row.get('expected_profit', 0)
        
    except Exception as e:
        log_error(f"Error creating paired orders sheet: {e}")


def create_buy_parameters_sheet(ws, theta: float, p: float, fit_metrics: Dict, 
                               budget: float, current_price: float, depths: np.ndarray) -> None:
    """Create buy-side parameters sheet."""
    try:
        # Title
        ws['A1'] = 'BUY-SIDE ANALYSIS'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Parameters
        row = 3
        ws[f'A{row}'] = 'Weibull Parameters'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Theta (scale):'
        ws[f'B{row}'] = f'{theta:.4f}'
        row += 1
        
        ws[f'A{row}'] = 'P (shape):'
        ws[f'B{row}'] = f'{p:.4f}'
        row += 1
        
        # Fit quality
        row += 1
        ws[f'A{row}'] = 'Fit Quality'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'R²:'
        ws[f'B{row}'] = f'{fit_metrics.get("r_squared", 0):.4f}'
        row += 1
        
        ws[f'A{row}'] = 'RMSE:'
        ws[f'B{row}'] = f'{fit_metrics.get("rmse", 0):.6f}'
        row += 1
        
        ws[f'A{row}'] = 'Data Points:'
        ws[f'B{row}'] = fit_metrics.get('n_points', 0)
        row += 1
        
        # Depth analysis
        row += 1
        ws[f'A{row}'] = 'Depth Analysis'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Min Depth:'
        ws[f'B{row}'] = f'{np.min(depths):.3f}%'
        row += 1
        
        ws[f'A{row}'] = 'Max Depth:'
        ws[f'B{row}'] = f'{np.max(depths):.3f}%'
        row += 1
        
        ws[f'A{row}'] = 'Depth Range:'
        ws[f'B{row}'] = f'{np.max(depths) - np.min(depths):.3f}%'
        
    except Exception as e:
        log_error(f"Error creating buy parameters sheet: {e}")


def create_sell_parameters_sheet(ws, theta_sell: float, p_sell: float, fit_metrics_sell: Dict, 
                                sell_depths: np.ndarray, actual_profits: np.ndarray) -> None:
    """Create sell-side parameters sheet."""
    try:
        # Title
        ws['A1'] = 'SELL-SIDE ANALYSIS'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Parameters
        row = 3
        ws[f'A{row}'] = 'Weibull Parameters'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Theta (scale):'
        ws[f'B{row}'] = f'{theta_sell:.4f}'
        row += 1
        
        ws[f'A{row}'] = 'P (shape):'
        ws[f'B{row}'] = f'{p_sell:.4f}'
        row += 1
        
        # Fit quality
        row += 1
        ws[f'A{row}'] = 'Fit Quality'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'R²:'
        ws[f'B{row}'] = f'{fit_metrics_sell.get("r_squared", 0):.4f}'
        row += 1
        
        ws[f'A{row}'] = 'RMSE:'
        ws[f'B{row}'] = f'{fit_metrics_sell.get("rmse", 0):.6f}'
        row += 1
        
        ws[f'A{row}'] = 'Data Points:'
        ws[f'B{row}'] = fit_metrics_sell.get('n_points', 0)
        row += 1
        
        # Profit analysis
        row += 1
        ws[f'A{row}'] = 'Profit Analysis'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Min Profit:'
        ws[f'B{row}'] = f'{np.min(actual_profits):.2f}%'
        row += 1
        
        ws[f'A{row}'] = 'Max Profit:'
        ws[f'B{row}'] = f'{np.max(actual_profits):.2f}%'
        row += 1
        
        ws[f'A{row}'] = 'Average Profit:'
        ws[f'B{row}'] = f'{np.mean(actual_profits):.2f}%'
        
    except Exception as e:
        log_error(f"Error creating sell parameters sheet: {e}")


def create_calculations_sheet(ws, depths: np.ndarray, allocations: np.ndarray, theta: float, p: float) -> None:
    """Create size calculations sheet."""
    try:
        # Title
        ws['A1'] = 'SIZE CALCULATIONS'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ['Rung', 'Depth (%)', 'Allocation ($)', 'Touch Probability', 'Expected Return']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        for i, (depth, allocation) in enumerate(zip(depths, allocations)):
            row = i + 4
            touch_prob = np.exp(-(depth / theta) ** p)
            expected_return = depth * touch_prob
            
            ws.cell(row=row, column=1).value = i + 1
            ws.cell(row=row, column=2).value = f'{depth:.3f}'
            ws.cell(row=row, column=3).value = f'{allocation:.2f}'
            ws.cell(row=row, column=4).value = f'{touch_prob:.4f}'
            ws.cell(row=row, column=5).value = f'{expected_return:.4f}'
        
    except Exception as e:
        log_error(f"Error creating calculations sheet: {e}")


def create_profit_sheet(ws, paired_orders_df: pd.DataFrame) -> None:
    """Create profit analysis sheet."""
    try:
        # Title
        ws['A1'] = 'PROFIT ANALYSIS'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Summary
        row = 3
        ws[f'A{row}'] = 'Profit Summary'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        if 'profit_pct' in paired_orders_df.columns:
            ws[f'A{row}'] = 'Total Profit Potential:'
            ws[f'B{row}'] = f'{paired_orders_df["profit_pct"].sum():.2f}%'
            row += 1
            
            ws[f'A{row}'] = 'Average Profit per Pair:'
            ws[f'B{row}'] = f'{paired_orders_df["profit_pct"].mean():.2f}%'
            row += 1
            
            ws[f'A{row}'] = 'Min Profit per Pair:'
            ws[f'B{row}'] = f'{paired_orders_df["profit_pct"].min():.2f}%'
            row += 1
            
            ws[f'A{row}'] = 'Max Profit per Pair:'
            ws[f'B{row}'] = f'{paired_orders_df["profit_pct"].max():.2f}%'
        
    except Exception as e:
        log_error(f"Error creating profit sheet: {e}")


def create_probability_sheet(ws, depths: np.ndarray, theta: float, p: float) -> None:
    """Create touch probability analysis sheet."""
    try:
        # Title
        ws['A1'] = 'TOUCH PROBABILITY ANALYSIS'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ['Depth (%)', 'Touch Probability', 'Expected Timeframe (hours)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        for i, depth in enumerate(depths):
            row = i + 4
            touch_prob = np.exp(-(depth / theta) ** p)
            expected_timeframe = 1.0 / touch_prob if touch_prob > 0 else np.inf
            
            ws.cell(row=row, column=1).value = f'{depth:.3f}'
            ws.cell(row=row, column=2).value = f'{touch_prob:.4f}'
            ws.cell(row=row, column=3).value = f'{expected_timeframe:.1f}' if expected_timeframe != np.inf else '∞'
        
    except Exception as e:
        log_error(f"Error creating probability sheet: {e}")


def create_scenario_analysis_sheet(ws, scenarios_df: pd.DataFrame) -> None:
    """Create scenario analysis sheet."""
    try:
        # Title
        ws['A1'] = 'SCENARIO ANALYSIS'
        ws['A1'].font = Font(size=14, bold=True)
        
        if scenarios_df.empty:
            return
        
        # Add data
        for row_idx, (_, row) in enumerate(scenarios_df.iterrows(), 3):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                if col_idx == 1:  # First column
                    ws.cell(row=row_idx, column=col_idx).value = col_name
                    ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)
                else:
                    ws.cell(row=row_idx, column=col_idx).value = value
        
    except Exception as e:
        log_error(f"Error creating scenario analysis sheet: {e}")


def create_sensitivity_analysis_sheet(ws, rung_sensitivity_df: pd.DataFrame, depth_sensitivity_df: pd.DataFrame) -> None:
    """Create sensitivity analysis sheet."""
    try:
        # Title
        ws['A1'] = 'SENSITIVITY ANALYSIS'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Rung sensitivity
        if rung_sensitivity_df is not None and not rung_sensitivity_df.empty:
            row = 3
            ws[f'A{row}'] = 'Rung Sensitivity'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for _, row_data in rung_sensitivity_df.iterrows():
                ws[f'A{row}'] = f'{row_data["num_rungs"]} rungs:'
                ws[f'B{row}'] = f'{row_data.get("capital_efficiency", 0):.4f} efficiency'
                row += 1
        
    except Exception as e:
        log_error(f"Error creating sensitivity analysis sheet: {e}")


def create_combined_sensitivity_sheet(ws, combined_sensitivity_df: pd.DataFrame) -> None:
    """Create combined sensitivity sheet."""
    try:
        # Title
        ws['A1'] = 'COMBINED SENSITIVITY ANALYSIS'
        ws['A1'].font = Font(size=14, bold=True)
        
        if combined_sensitivity_df is None or combined_sensitivity_df.empty:
            return
        
        # Add data
        for row_idx, (_, row) in enumerate(combined_sensitivity_df.iterrows(), 3):
            ws.cell(row=row_idx, column=1).value = f'{row["num_rungs"]} rungs, {row["depth_range_name"]}'
            ws.cell(row=row_idx, column=2).value = f'{row.get("capital_efficiency", 0):.4f}'
        
    except Exception as e:
        log_error(f"Error creating combined sensitivity sheet: {e}")


if __name__ == "__main__":
    # Test with sample data
    import numpy as np
    import pandas as pd
    
    # Sample data
    depths = np.array([0.5, 1.0, 1.5, 2.0, 2.5])
    empirical_probs = np.array([0.8, 0.6, 0.4, 0.2, 0.1])
    theta, p = 2.0, 1.5
    fit_metrics = {'r_squared': 0.95, 'rmse': 0.05, 'n_points': 100}
    
    depths_upward = np.array([0.5, 1.0, 1.5, 2.0, 2.5])
    empirical_probs_upward = np.array([0.7, 0.5, 0.3, 0.15, 0.05])
    theta_sell, p_sell = 2.2, 1.3
    fit_metrics_sell = {'r_squared': 0.92, 'rmse': 0.06, 'n_points': 100}
    
    allocations = np.array([100, 120, 140, 160, 180])
    
    paired_orders_df = pd.DataFrame({
        'rung': range(1, 6),
        'buy_price': [99.5, 99.0, 98.5, 98.0, 97.5],
        'sell_price': [100.5, 101.0, 101.5, 102.0, 102.5],
        'buy_qty': [1.0, 1.2, 1.4, 1.6, 1.8],
        'sell_qty': [1.0, 1.2, 1.4, 1.6, 1.8],
        'profit_pct': [1.0, 2.0, 3.0, 4.0, 5.0],
        'profit_usd': [1.0, 2.4, 4.2, 6.4, 9.0]
    })
    
    scenarios_df = pd.DataFrame({
        'profit_target_pct': [1.0, 2.0, 3.0],
        'expected_monthly_profit': [100, 200, 300],
        'expected_timeframe_hours': [24, 48, 72]
    })
    
    optimal_scenario = {'profit_target_pct': 2.0, 'expected_monthly_profit': 200}
    
    log_info("=== VISUALIZATION TEST ===")
    create_all_visualizations(depths, empirical_probs, theta, p, fit_metrics,
                            depths, allocations, paired_orders_df,
                            depths_upward, empirical_probs_upward, theta_sell, p_sell, fit_metrics_sell,
                            depths_upward, np.array([1.0, 2.0, 3.0, 4.0, 5.0]),
                            scenarios_df, optimal_scenario)
    
    log_info("\n=== EXCEL EXPORT TEST ===")
    create_excel_workbook(paired_orders_df, depths, allocations, theta, p, fit_metrics,
                        10000.0, 100.0, theta_sell, p_sell, fit_metrics_sell,
                        depths_upward, np.array([1.0, 2.0, 3.0, 4.0, 5.0]),
                        scenarios_df)
