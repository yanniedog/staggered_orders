"""
Excel export module for ladder analysis and order specifications.
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict
import os


def create_output_dir():
    """Create output directory if it doesn't exist"""
    try:
        if not os.path.exists('output'):
            os.makedirs('output')
    except Exception as e:
        print(f"Warning: Could not create output directory: {e}")


def create_excel_workbook(paired_orders_df: pd.DataFrame, depths: np.ndarray, 
                         allocations: np.ndarray, theta: float, p: float,
                         fit_metrics: Dict, budget: float, current_price: float,
                         theta_sell: float, p_sell: float, fit_metrics_sell: Dict,
                         sell_depths: np.ndarray, actual_profits: np.ndarray,
                         scenarios_df: pd.DataFrame = None,
                         rung_sensitivity_df: pd.DataFrame = None,
                         depth_sensitivity_df: pd.DataFrame = None,
                         combined_sensitivity_df: pd.DataFrame = None) -> None:
    """
    Create comprehensive Excel workbook with paired order specs, analysis, and scenario data.
    """
    try:
        create_output_dir()
        
        # Validate inputs
        if paired_orders_df is None or paired_orders_df.empty:
            print("Warning: Empty paired orders DataFrame provided to Excel export")
            return
        
        if np.any(np.isnan(depths)) or np.any(np.isnan(allocations)):
            print("Warning: NaN values in depths or allocations")
            return
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Executive Summary
        ws_executive = wb.create_sheet("Executive Summary")
        create_executive_summary_sheet(ws_executive, paired_orders_df, fit_metrics, fit_metrics_sell, 
                                     scenarios_df, budget, current_price)
        
        # Sheet 2: Sanity Checks & Validation
        ws_sanity = wb.create_sheet("Sanity Checks")
        create_sanity_checks_sheet(ws_sanity, paired_orders_df, fit_metrics, fit_metrics_sell, 
                                 scenarios_df, budget, current_price)
        
        # Sheet 3: Paired Order Specifications
        ws_paired = wb.create_sheet("Paired Orders")
        create_paired_orders_sheet(ws_paired, paired_orders_df)
        
        # Sheet 4: Buy-Side Parameters & Analysis
        ws_buy_params = wb.create_sheet("Buy-Side Analysis")
        create_buy_parameters_sheet(ws_buy_params, theta, p, fit_metrics, budget, current_price, depths)
        
        # Sheet 5: Sell-Side Parameters & Analysis
        ws_sell_params = wb.create_sheet("Sell-Side Analysis")
        create_sell_parameters_sheet(ws_sell_params, theta_sell, p_sell, fit_metrics_sell, sell_depths, actual_profits)
        
        # Sheet 6: Size Calculations
        ws_calc = wb.create_sheet("Size Calculations")
        create_calculations_sheet(ws_calc, depths, allocations, theta, p)
        
        # Sheet 7: Profit Analysis
        ws_profit = wb.create_sheet("Profit Analysis")
        create_profit_sheet(ws_profit, paired_orders_df)
        
        # Sheet 8: Touch Probability Analysis
        ws_prob = wb.create_sheet("Touch Probability")
        create_probability_sheet(ws_prob, depths, theta, p)
        
        # New scenario analysis sheets
        if scenarios_df is not None:
            # Sheet 9: Scenario Analysis
            ws_scenarios = wb.create_sheet("Scenario Analysis")
            create_scenario_analysis_sheet(ws_scenarios, scenarios_df)
            
            # Sheet 10: Sensitivity Analysis
            ws_sensitivity = wb.create_sheet("Sensitivity Analysis")
            create_sensitivity_analysis_sheet(ws_sensitivity, rung_sensitivity_df, depth_sensitivity_df)
            
            # Sheet 11: Combined Sensitivity
            ws_combined = wb.create_sheet("Combined Sensitivity")
            create_combined_sensitivity_sheet(ws_combined, combined_sensitivity_df)
            
            # Sheet 12: Recommendations
            ws_recommendations = wb.create_sheet("Recommendations")
            create_recommendations_sheet(ws_recommendations, scenarios_df, rung_sensitivity_df, depth_sensitivity_df)
        
        # Save workbook
        filename = 'output/ladder_report.xlsx'
        wb.save(filename)
        print(f"Excel workbook saved to {filename}")
        
    except Exception as e:
        print(f"Error creating Excel workbook: {e}")
        import traceback
        traceback.print_exc()


def create_paired_orders_sheet(ws, paired_orders_df: pd.DataFrame):
    """Create paired order specifications sheet"""
    # Headers
    headers = ['Rung', 'Buy Depth (%)', 'Buy Price ($)', 'Buy Qty', 'Buy Notional ($)',
               'Sell Depth (%)', 'Sell Price ($)', 'Sell Qty', 'Sell Notional ($)',
               'Profit (%)', 'Profit ($)', 'Target Profit (%)']
    
    # Style headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, (_, row) in enumerate(paired_orders_df.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row['rung'])
        ws.cell(row=row_idx, column=2, value=round(row['buy_depth_pct'], 3))
        ws.cell(row=row_idx, column=3, value=round(row['buy_price'], 2))
        ws.cell(row=row_idx, column=4, value=round(row['buy_qty'], 3))
        ws.cell(row=row_idx, column=5, value=round(row['buy_notional'], 2))
        ws.cell(row=row_idx, column=6, value=round(row['sell_depth_pct'], 3))
        ws.cell(row=row_idx, column=7, value=round(row['sell_price'], 2))
        ws.cell(row=row_idx, column=8, value=round(row['sell_qty'], 3))
        ws.cell(row=row_idx, column=9, value=round(row['sell_notional'], 2))
        ws.cell(row=row_idx, column=10, value=round(row['profit_pct'], 2))
        ws.cell(row=row_idx, column=11, value=round(row['profit_usd'], 2))
        ws.cell(row=row_idx, column=12, value=round(row['target_profit_pct'], 2))
    
    # Format columns
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col].width = 12
    
    # Add summary with proper row references
    summary_row = len(paired_orders_df) + 3
    last_data_row = len(paired_orders_df) + 1
    ws.cell(row=summary_row, column=1, value="TOTALS:").font = Font(bold=True)
    # Column E: buy_notional, Column I: sell_notional, Column K: profit_usd
    ws.cell(row=summary_row, column=5, value=f"=SUM(E2:E{last_data_row})").font = Font(bold=True)
    ws.cell(row=summary_row, column=9, value=f"=SUM(I2:I{last_data_row})").font = Font(bold=True)
    ws.cell(row=summary_row, column=11, value=f"=SUM(K2:K{last_data_row})").font = Font(bold=True)


def create_buy_parameters_sheet(ws, theta: float, p: float, fit_metrics: Dict,
                               budget: float, current_price: float, depths: np.ndarray):
    """Create buy-side parameters and analysis sheet"""
    # Parameters section
    ws.cell(row=1, column=1, value="BUY-SIDE PARAMETERS").font = Font(bold=True, size=14)
    
    params = [
        ("Weibull Theta (scale)", theta),
        ("Weibull P (shape)", p),
        ("Fit R²", fit_metrics['r_squared']),
        ("Fit RMSE", fit_metrics['rmse']),
        ("Budget (USD)", budget),
        ("Current Price", current_price),
        ("Number of Rungs", len(depths)),
        ("Depth Range", f"{depths[0]:.3f}% - {depths[-1]:.3f}%"),
        ("Total Allocation", f"=SUM('Paired Orders'!E:E)"),
    ]
    
    for i, (param, value) in enumerate(params, 2):
        ws.cell(row=i, column=1, value=param).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)


def create_sell_parameters_sheet(ws, theta_sell: float, p_sell: float, fit_metrics_sell: Dict,
                                sell_depths: np.ndarray, actual_profits: np.ndarray):
    """Create sell-side parameters and analysis sheet"""
    # Parameters section
    ws.cell(row=1, column=1, value="SELL-SIDE PARAMETERS").font = Font(bold=True, size=14)
    
    params = [
        ("Weibull Theta (scale)", theta_sell),
        ("Weibull P (shape)", p_sell),
        ("Fit R²", fit_metrics_sell['r_squared']),
        ("Fit RMSE", fit_metrics_sell['rmse']),
        ("Number of Sell Rungs", len(sell_depths)),
        ("Sell Depth Range", f"{sell_depths[0]:.3f}% - {sell_depths[-1]:.3f}%"),
        ("Total Expected Profit", f"=SUM('Paired Orders'!K:K)"),
        ("Average Profit", f"=AVERAGE('Paired Orders'!J:J)"),
        ("Min Profit", f"=MIN('Paired Orders'!J:J)"),
        ("Max Profit", f"=MAX('Paired Orders'!J:J)"),
    ]
    
    for i, (param, value) in enumerate(params, 2):
        ws.cell(row=i, column=1, value=param).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)


def create_profit_sheet(ws, paired_orders_df: pd.DataFrame):
    """Create profit analysis sheet"""
    # Headers
    headers = ['Rung', 'Buy Depth (%)', 'Sell Depth (%)', 'Profit (%)', 'Profit ($)', 
               'Target Profit (%)', 'Profit Ratio', 'Buy Notional ($)', 'Sell Notional ($)']
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, (_, row) in enumerate(paired_orders_df.iterrows(), 2):
        profit_ratio = row['profit_pct'] / row['target_profit_pct'] if row['target_profit_pct'] > 0 else 0
        
        ws.cell(row=row_idx, column=1, value=row['rung'])
        ws.cell(row=row_idx, column=2, value=round(row['buy_depth_pct'], 3))
        ws.cell(row=row_idx, column=3, value=round(row['sell_depth_pct'], 3))
        ws.cell(row=row_idx, column=4, value=round(row['profit_pct'], 2))
        ws.cell(row=row_idx, column=5, value=round(row['profit_usd'], 2))
        ws.cell(row=row_idx, column=6, value=round(row['target_profit_pct'], 2))
        ws.cell(row=row_idx, column=7, value=round(profit_ratio, 2))
        ws.cell(row=row_idx, column=8, value=round(row['buy_notional'], 2))
        ws.cell(row=row_idx, column=9, value=round(row['sell_notional'], 2))
    
    # Format columns
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 15
    
    # Add summary statistics
    summary_start = len(paired_orders_df) + 3
    ws.cell(row=summary_start, column=1, value="PROFIT SUMMARY").font = Font(bold=True, size=12)
    
    stats = [
        ("Total Expected Profit", f"=SUM(E2:E{len(paired_orders_df)+1})"),
        ("Average Profit %", f"=AVERAGE(D2:D{len(paired_orders_df)+1})"),
        ("Min Profit %", f"=MIN(D2:D{len(paired_orders_df)+1})"),
        ("Max Profit %", f"=MAX(D2:D{len(paired_orders_df)+1})"),
        ("Profit Std Dev", f"=STDEV(D2:D{len(paired_orders_df)+1})"),
        ("Average Profit Ratio", f"=AVERAGE(G2:G{len(paired_orders_df)+1})"),
    ]
    
    for i, (stat, value) in enumerate(stats, summary_start + 1):
        ws.cell(row=i, column=1, value=stat).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)


def create_orders_sheet(ws, orders_df: pd.DataFrame):
    """Create order specifications sheet"""
    # Headers
    headers = ['Rung', 'Depth (%)', 'Limit Price ($)', 'Quantity', 'Notional ($)', 'Allocation ($)']
    
    # Style headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, (_, row) in enumerate(orders_df.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row['rung'])
        ws.cell(row=row_idx, column=2, value=round(row['depth_pct'], 3))
        ws.cell(row=row_idx, column=3, value=round(row['limit_price'], 2))
        ws.cell(row=row_idx, column=4, value=round(row['quantity'], 3))
        ws.cell(row=row_idx, column=5, value=round(row['notional'], 2))
        ws.cell(row=row_idx, column=6, value=round(row['allocation'], 2))
    
    # Format columns
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Add summary with proper row references
    summary_row = len(orders_df) + 3
    last_data_row = len(orders_df) + 1
    ws.cell(row=summary_row, column=1, value="TOTAL:").font = Font(bold=True)
    # Column E: notional, Column F: allocation
    ws.cell(row=summary_row, column=5, value=f"=SUM(E2:E{last_data_row})").font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=f"=SUM(F2:F{last_data_row})").font = Font(bold=True)


def create_parameters_sheet(ws, theta: float, p: float, fit_metrics: Dict,
                           budget: float, current_price: float, depths: np.ndarray):
    """Create parameters and analysis sheet"""
    # Parameters section
    ws.cell(row=1, column=1, value="PARAMETERS").font = Font(bold=True, size=14)
    
    params = [
        ("Weibull Theta (scale)", theta),
        ("Weibull P (shape)", p),
        ("Fit R²", fit_metrics['r_squared']),
        ("Fit RMSE", fit_metrics['rmse']),
        ("Budget (USD)", budget),
        ("Current Price", current_price),
        ("Number of Rungs", len(depths)),
        ("Depth Range", f"{depths[0]:.3f}% - {depths[-1]:.3f}%"),
        ("Total Allocation", f"=SUM('Order Specifications'!F:F)"),
        ("Total Notional", f"=SUM('Order Specifications'!E:E)")
    ]
    
    for i, (param, value) in enumerate(params, 2):
        ws.cell(row=i, column=1, value=param).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    
    # Analysis section
    analysis_start = len(params) + 4
    ws.cell(row=analysis_start, column=1, value="ANALYSIS").font = Font(bold=True, size=14)
    
    # Calculate some analysis metrics
    d_min, d_max = depths[0], depths[-1]
    alpha = max(1.0, p * (d_max / theta) ** p)
    
    analysis = [
        ("Alpha (monotonicity)", alpha),
        ("Depth Range Width", f"{d_max - d_min:.3f}%"),
        ("Average Depth", f"{np.mean(depths):.3f}%"),
        ("Median Depth", f"{np.median(depths):.3f}%"),
        ("Shallowest Touch Prob", f"{np.exp(-(d_min/theta)**p):.4f}"),
        ("Deepest Touch Prob", f"{np.exp(-(d_max/theta)**p):.4f}")
    ]
    
    for i, (metric, value) in enumerate(analysis, analysis_start + 1):
        ws.cell(row=i, column=1, value=metric).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    
    # Format columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def create_calculations_sheet(ws, depths: np.ndarray, allocations: np.ndarray,
                             theta: float, p: float):
    """Create size calculations sheet with formulas"""
    # Headers
    headers = ['Rung', 'Depth (%)', 'Touch Prob', 'Weight', 'Allocation ($)', 'Alpha']
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Calculate alpha
    alpha = max(1.0, p * (depths[-1] / theta) ** p)
    
    # Add data and formulas
    for i, (depth, allocation) in enumerate(zip(depths, allocations), 2):
        touch_prob = np.exp(-(depth / theta) ** p)
        weight = (depth ** alpha) * np.exp(-(depth / theta) ** p)
        
        ws.cell(row=i, column=1, value=i-1)  # Rung number
        ws.cell(row=i, column=2, value=round(depth, 3))
        ws.cell(row=i, column=3, value=round(touch_prob, 6))
        ws.cell(row=i, column=4, value=round(weight, 6))
        ws.cell(row=i, column=5, value=round(allocation, 2))
        ws.cell(row=i, column=6, value=round(alpha, 3))
    
    # Add formulas for recalculation
    formula_start = len(depths) + 3
    ws.cell(row=formula_start, column=1, value="FORMULAS FOR RECALCULATION").font = Font(bold=True, size=12)
    
    formulas = [
        ("Touch Probability", "=EXP(-(B2/theta)^p)"),
        ("Weight Function", "=B2^alpha*EXP(-(B2/theta)^p)"),
        ("Normalized Weight", "=D2/SUM(D:D)"),
        ("New Allocation", "=budget*E2"),
        ("Alpha Calculation", "=MAX(1,p*(max_depth/theta)^p)")
    ]
    
    for i, (desc, formula) in enumerate(formulas, formula_start + 1):
        ws.cell(row=i, column=1, value=desc).font = Font(bold=True)
        ws.cell(row=i, column=2, value=formula)
    
    # Format columns
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18


def create_probability_sheet(ws, depths: np.ndarray, theta: float, p: float):
    """Create touch probability analysis sheet"""
    # Headers
    headers = ['Depth (%)', 'Touch Probability', 'Expected Return', 'Cumulative Prob']
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Generate detailed probability curve
    detail_depths = np.linspace(0.1, depths[-1] * 1.5, 100)
    touch_probs = np.exp(-(detail_depths / theta) ** p)
    expected_returns = detail_depths * touch_probs
    cumulative_probs = 1 - np.exp(-(detail_depths / theta) ** p)
    
    # Add data
    for i, (depth, prob, exp_ret, cum_prob) in enumerate(zip(detail_depths, touch_probs, 
                                                             expected_returns, cumulative_probs), 2):
        ws.cell(row=i, column=1, value=round(depth, 3))
        ws.cell(row=i, column=2, value=round(prob, 6))
        ws.cell(row=i, column=3, value=round(exp_ret, 6))
        ws.cell(row=i, column=4, value=round(cum_prob, 6))
    
    # Add summary statistics
    summary_start = len(detail_depths) + 3
    ws.cell(row=summary_start, column=1, value="SUMMARY STATISTICS").font = Font(bold=True, size=12)
    
    stats = [
        ("Max Expected Return", f"={max(expected_returns):.6f}"),
        ("Depth at Max Return", f"={detail_depths[np.argmax(expected_returns)]:.3f}%"),
        ("Mean Touch Probability", f"={np.mean(touch_probs):.6f}"),
        ("Median Touch Probability", f"={np.median(touch_probs):.6f}")
    ]
    
    for i, (stat, value) in enumerate(stats, summary_start + 1):
        ws.cell(row=i, column=1, value=stat).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    
    # Format columns
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 20


def create_scenario_analysis_sheet(ws, scenarios_df: pd.DataFrame):
    """Create scenario analysis sheet"""
    # Headers
    headers = ['Rank', 'Profit Target (%)', 'Num Rungs', 'Buy Depth Min (%)', 'Buy Depth Max (%)',
               'Sell Depth Min (%)', 'Sell Depth Max (%)', 'Expected Profit/Dollar', 
               'Expected Timeframe (h)', 'Capital Efficiency', 'Risk-Adjusted Return',
               'Joint Touch Prob', 'Avg Buy Touch Prob', 'Avg Sell Touch Prob']
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, (_, row) in enumerate(scenarios_df.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=int(row['rank']))
        ws.cell(row=row_idx, column=2, value=round(row['profit_target_pct'], 1))
        ws.cell(row=row_idx, column=3, value=int(row['num_rungs']))
        ws.cell(row=row_idx, column=4, value=round(row['buy_depth_min'], 3))
        ws.cell(row=row_idx, column=5, value=round(row['buy_depth_max'], 3))
        ws.cell(row=row_idx, column=6, value=round(row['sell_depth_min'], 3))
        ws.cell(row=row_idx, column=7, value=round(row['sell_depth_max'], 3))
        ws.cell(row=row_idx, column=8, value=round(row['expected_profit_per_dollar'], 4))
        ws.cell(row=row_idx, column=9, value=round(row['expected_timeframe_hours'], 1))
        ws.cell(row=row_idx, column=10, value=round(row['expected_profit_per_dollar'], 4))
        ws.cell(row=row_idx, column=11, value=round(row['sharpe_ratio'], 3))
        ws.cell(row=row_idx, column=12, value=round(row['joint_touch_prob'], 3))
        ws.cell(row=row_idx, column=13, value=round(row['avg_buy_touch_prob'], 3))
        ws.cell(row=row_idx, column=14, value=round(row['avg_sell_touch_prob'], 3))
    
    # Format columns
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 15
    
    # Add summary
    summary_row = len(scenarios_df) + 3
    ws.cell(row=summary_row, column=1, value="SCENARIO SUMMARY").font = Font(bold=True, size=12)
    
    summary_stats = [
        ("Total Scenarios", len(scenarios_df)),
        ("Best Expected Value", f"=MAX(H2:H{len(scenarios_df)+1})"),
        ("Best Timeframe", f"=MIN(I2:I{len(scenarios_df)+1})"),
        ("Best Risk-Adjusted", f"=MAX(K2:K{len(scenarios_df)+1})"),
        ("Average Profit Target", f"=AVERAGE(B2:B{len(scenarios_df)+1})"),
        ("Average Rungs", f"=AVERAGE(C2:C{len(scenarios_df)+1})")
    ]
    
    for i, (stat, value) in enumerate(summary_stats, summary_row + 1):
        ws.cell(row=i, column=1, value=stat).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)


def create_sensitivity_analysis_sheet(ws, rung_sensitivity_df: pd.DataFrame, depth_sensitivity_df: pd.DataFrame):
    """Create sensitivity analysis sheet"""
    # Rung sensitivity section
    ws.cell(row=1, column=1, value="RUNG SENSITIVITY ANALYSIS").font = Font(bold=True, size=14)
    
    rung_headers = ['Num Rungs', 'Expected Profit/Dollar', 'Expected Timeframe (h)', 
                   'Capital Efficiency', 'Risk-Adjusted Return', 'Allocation Ratio']
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col, header in enumerate(rung_headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    if rung_sensitivity_df is not None:
        for row_idx, (_, row) in enumerate(rung_sensitivity_df.iterrows(), 3):
            ws.cell(row=row_idx, column=1, value=int(row['num_rungs']))
            ws.cell(row=row_idx, column=2, value=round(row['expected_profit_per_dollar'], 4))
            ws.cell(row=row_idx, column=3, value=round(row['expected_timeframe_hours'], 1))
            ws.cell(row=row_idx, column=4, value=round(row['capital_efficiency'], 4))
            ws.cell(row=row_idx, column=5, value=round(row['risk_adjusted_return'], 3))
            ws.cell(row=row_idx, column=6, value=round(row['allocation_ratio'], 2))
    
    # Depth sensitivity section
    depth_start = 3 + (len(rung_sensitivity_df) if rung_sensitivity_df is not None else 0) + 2
    ws.cell(row=depth_start, column=1, value="DEPTH SENSITIVITY ANALYSIS").font = Font(bold=True, size=14)
    
    depth_headers = ['Strategy', 'Capital Efficiency', 'Expected Timeframe (h)', 
                    'Expected Fills', 'Depth Range (%)', 'Any Fill Probability']
    
    for col, header in enumerate(depth_headers, 1):
        cell = ws.cell(row=depth_start + 1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    if depth_sensitivity_df is not None:
        for row_idx, (_, row) in enumerate(depth_sensitivity_df.iterrows(), depth_start + 2):
            ws.cell(row=row_idx, column=1, value=row['strategy'])
            ws.cell(row=row_idx, column=2, value=round(row['capital_efficiency'], 4))
            ws.cell(row=row_idx, column=3, value=round(row['expected_timeframe_hours'], 1))
            ws.cell(row=row_idx, column=4, value=round(row['expected_fills'], 2))
            ws.cell(row=row_idx, column=5, value=round(row['depth_range'], 2))
            ws.cell(row=row_idx, column=6, value=round(row['any_fill_probability'], 3))
    
    # Format columns
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18


def create_combined_sensitivity_sheet(ws, combined_sensitivity_df: pd.DataFrame):
    """Create combined sensitivity analysis sheet"""
    # Headers
    headers = ['Num Rungs', 'Strategy', 'Profit Target (%)', 'Buy Depth Min (%)', 'Buy Depth Max (%)',
               'Depth Range (%)', 'Expected Profit/Dollar', 'Expected Timeframe (h)',
               'Capital Efficiency', 'Risk Efficiency', 'Combined Score', 'Allocation Ratio']
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    if combined_sensitivity_df is not None:
        # Add data
        for row_idx, (_, row) in enumerate(combined_sensitivity_df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=int(row['num_rungs']))
            ws.cell(row=row_idx, column=2, value=row['strategy'])
            ws.cell(row=row_idx, column=3, value=round(row['profit_target_pct'], 1))
            ws.cell(row=row_idx, column=4, value=round(row['buy_depth_min'], 3))
            ws.cell(row=row_idx, column=5, value=round(row['buy_depth_max'], 3))
            ws.cell(row=row_idx, column=6, value=round(row['depth_range'], 2))
            ws.cell(row=row_idx, column=7, value=round(row['expected_profit_per_dollar'], 4))
            ws.cell(row=row_idx, column=8, value=round(row['expected_timeframe_hours'], 1))
            ws.cell(row=row_idx, column=9, value=round(row['capital_efficiency'], 4))
            ws.cell(row=row_idx, column=10, value=round(row['risk_efficiency'], 3))
            ws.cell(row=row_idx, column=11, value=round(row['combined_score'], 4))
            ws.cell(row=row_idx, column=12, value=round(row['allocation_ratio'], 2))
    
    # Format columns
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col].width = 15
    
    # Add summary
    if combined_sensitivity_df is not None:
        summary_row = len(combined_sensitivity_df) + 3
        ws.cell(row=summary_row, column=1, value="COMBINED SENSITIVITY SUMMARY").font = Font(bold=True, size=12)
        
        summary_stats = [
            ("Best Combined Score", f"=MAX(K2:K{len(combined_sensitivity_df)+1})"),
            ("Best Capital Efficiency", f"=MAX(I2:I{len(combined_sensitivity_df)+1})"),
            ("Best Risk Efficiency", f"=MAX(J2:J{len(combined_sensitivity_df)+1})"),
            ("Fastest Timeframe", f"=MIN(H2:H{len(combined_sensitivity_df)+1})"),
            ("Total Configurations", len(combined_sensitivity_df))
        ]
        
        for i, (stat, value) in enumerate(summary_stats, summary_row + 1):
            ws.cell(row=i, column=1, value=stat).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)


def create_recommendations_sheet(ws, scenarios_df: pd.DataFrame, rung_sensitivity_df: pd.DataFrame, depth_sensitivity_df: pd.DataFrame):
    """Create recommendations sheet"""
    ws.cell(row=1, column=1, value="RECOMMENDATIONS & INSIGHTS").font = Font(bold=True, size=16)
    
    # Top scenarios
    ws.cell(row=3, column=1, value="TOP 3 PROFIT SCENARIOS").font = Font(bold=True, size=14)
    
    top_scenarios = scenarios_df.head(3)
    for i, (_, row) in enumerate(top_scenarios.iterrows(), 4):
        ws.cell(row=i, column=1, value=f"{i-3}. {row['profit_target_pct']:.1f}% Profit Target").font = Font(bold=True)
        ws.cell(row=i, column=2, value=f"{row['num_rungs']} rungs, {row['expected_timeframe_hours']:.1f}h timeframe")
        ws.cell(row=i+1, column=1, value=f"Expected Value: {row['expected_profit_per_dollar']:.4f}")
        ws.cell(row=i+1, column=2, value=f"Risk Score: {row['sharpe_ratio']:.3f}")
    
    # Optimal rung count
    rung_start = 12
    ws.cell(row=rung_start, column=1, value="OPTIMAL RUNG COUNT ANALYSIS").font = Font(bold=True, size=14)
    
    if rung_sensitivity_df is not None:
        best_rungs = rung_sensitivity_df.loc[rung_sensitivity_df['capital_efficiency'].idxmax()]
        ws.cell(row=rung_start+1, column=1, value="Recommended Rung Count:").font = Font(bold=True)
        ws.cell(row=rung_start+1, column=2, value=f"{int(best_rungs['num_rungs'])} rungs")
        ws.cell(row=rung_start+2, column=1, value="Reasoning:")
        ws.cell(row=rung_start+2, column=2, value=f"Maximizes capital efficiency ({best_rungs['capital_efficiency']:.4f})")
        ws.cell(row=rung_start+3, column=1, value="Expected Timeframe:")
        ws.cell(row=rung_start+3, column=2, value=f"{best_rungs['expected_timeframe_hours']:.1f} hours")
    
    # Best depth strategy
    depth_start = rung_start + 6
    ws.cell(row=depth_start, column=1, value="OPTIMAL DEPTH STRATEGY").font = Font(bold=True, size=14)
    
    if depth_sensitivity_df is not None:
        best_strategy = depth_sensitivity_df.loc[depth_sensitivity_df['capital_efficiency'].idxmax()]
        ws.cell(row=depth_start+1, column=1, value="Recommended Strategy:").font = Font(bold=True)
        ws.cell(row=depth_start+1, column=2, value=best_strategy['strategy'])
        ws.cell(row=depth_start+2, column=1, value="Capital Efficiency:")
        ws.cell(row=depth_start+2, column=2, value=f"{best_strategy['capital_efficiency']:.4f}")
        ws.cell(row=depth_start+3, column=1, value="Expected Fills:")
        ws.cell(row=depth_start+3, column=2, value=f"{best_strategy['expected_fills']:.1f}")
    
    # Risk considerations
    risk_start = depth_start + 6
    ws.cell(row=risk_start, column=1, value="RISK CONSIDERATIONS").font = Font(bold=True, size=14)
    
    ws.cell(row=risk_start+1, column=1, value="Key Risk Factors:")
    ws.cell(row=risk_start+2, column=1, value="• Timeframe uncertainty increases with deeper rungs")
    ws.cell(row=risk_start+3, column=1, value="• Higher profit targets require longer holding periods")
    ws.cell(row=risk_start+4, column=1, value="• Joint touch probability decreases with deeper strategies")
    ws.cell(row=risk_start+5, column=1, value="• Market regime changes can affect Weibull parameters")
    
    # Format columns
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40


def create_executive_summary_sheet(ws, paired_orders_df: pd.DataFrame, fit_metrics: Dict, 
                                 fit_metrics_sell: Dict, scenarios_df: pd.DataFrame,
                                 budget: float, current_price: float) -> None:
    """
    Create executive summary sheet with key metrics and insights.
    """
    # Title
    ws['A1'] = "EXECUTIVE SUMMARY"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    
    row = 3
    
    # Key Metrics Section
    ws[f'A{row}'] = "KEY METRICS"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    # Calculate key metrics
    total_buy_notional = paired_orders_df['buy_notional'].sum()
    total_sell_notional = paired_orders_df['sell_notional'].sum()
    total_expected_profit = paired_orders_df['profit_usd'].sum()
    avg_profit_pct = paired_orders_df['profit_pct'].mean()
    num_pairs = len(paired_orders_df)
    
    # Buy-side fit quality
    buy_quality = fit_metrics.get('fit_quality', 'unknown')
    buy_r_squared = fit_metrics.get('r_squared', 0)
    
    # Sell-side fit quality
    sell_quality = fit_metrics_sell.get('fit_quality', 'unknown')
    sell_r_squared = fit_metrics_sell.get('r_squared', 0)
    
    # Key metrics table
    metrics = [
        ("Total Budget", f"${budget:,.0f}"),
        ("Current Price", f"${current_price:.2f}"),
        ("Number of Pairs", f"{num_pairs}"),
        ("Total Buy Notional", f"${total_buy_notional:,.0f}"),
        ("Total Sell Notional", f"${total_sell_notional:,.0f}"),
        ("Total Expected Profit", f"${total_expected_profit:,.0f}"),
        ("Average Profit per Pair", f"{avg_profit_pct:.1f}%"),
        ("Buy-Side Fit Quality", f"{buy_quality} (R²={buy_r_squared:.3f})"),
        ("Sell-Side Fit Quality", f"{sell_quality} (R²={sell_r_squared:.3f})")
    ]
    
    for metric, value in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    row += 1
    
    # Strategy Overview
    ws[f'A{row}'] = "STRATEGY OVERVIEW"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    overview_text = [
        f"This ladder strategy deploys {num_pairs} buy-sell pairs across different market depths.",
        f"The strategy expects to generate ${total_expected_profit:,.0f} in total profit",
        f"with an average profit of {avg_profit_pct:.1f}% per successful pair.",
        "",
        f"Buy-side analysis shows {buy_quality} fit quality with R² = {buy_r_squared:.3f}.",
        f"Sell-side analysis shows {sell_quality} fit quality with R² = {sell_r_squared:.3f}.",
        "",
        "The strategy is designed to capture market wicks while maintaining",
        "reasonable profit targets and risk management."
    ]
    
    for text in overview_text:
        ws[f'A{row}'] = text
        row += 1
    
    row += 1
    
    # Risk Assessment
    ws[f'A{row}'] = "RISK ASSESSMENT"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    # Calculate risk metrics
    max_profit = paired_orders_df['profit_pct'].max()
    min_profit = paired_orders_df['profit_pct'].min()
    profit_range = max_profit - min_profit
    
    risk_text = [
        f"Profit range: {min_profit:.1f}% to {max_profit:.1f}% (range: {profit_range:.1f}%)",
        f"Maximum single position: ${paired_orders_df['buy_notional'].max():,.0f}",
        f"Average position size: ${paired_orders_df['buy_notional'].mean():,.0f}",
        "",
        "Risk Factors:",
        "• Market volatility may affect fill rates",
        "• Deeper rungs have lower probability but higher profit",
        "• Joint probability decreases with strategy complexity"
    ]
    
    for text in risk_text:
        ws[f'A{row}'] = text
        row += 1
    
    # Format columns
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def create_sanity_checks_sheet(ws, paired_orders_df: pd.DataFrame, fit_metrics: Dict,
                              fit_metrics_sell: Dict, scenarios_df: pd.DataFrame,
                              budget: float, current_price: float) -> None:
    """
    Create sanity checks sheet with validation results and warnings.
    """
    # Title
    ws['A1'] = "SANITY CHECKS & VALIDATION"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    
    row = 3
    
    # Validation Results
    ws[f'A{row}'] = "VALIDATION RESULTS"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    # Check 1: Profitability
    unprofitable_pairs = paired_orders_df[paired_orders_df['profit_pct'] <= 0]
    profitability_status = "PASS" if len(unprofitable_pairs) == 0 else "FAIL"
    profitability_color = "00FF00" if profitability_status == "PASS" else "FF0000"
    
    ws[f'A{row}'] = "1. Profitability Check"
    ws[f'B{row}'] = profitability_status
    ws[f'C{row}'] = f"{len(unprofitable_pairs)} unprofitable pairs" if len(unprofitable_pairs) > 0 else "All pairs profitable"
    ws[f'B{row}'].fill = PatternFill(start_color=profitability_color, end_color=profitability_color, fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
    row += 1
    
    # Check 2: Profit Range Reasonableness
    max_profit = paired_orders_df['profit_pct'].max()
    profit_range_status = "PASS" if max_profit <= 50 else "WARNING" if max_profit <= 100 else "FAIL"
    profit_range_color = "00FF00" if profit_range_status == "PASS" else "FFA500" if profit_range_status == "WARNING" else "FF0000"
    
    ws[f'A{row}'] = "2. Profit Range Check"
    ws[f'B{row}'] = profit_range_status
    ws[f'C{row}'] = f"Max profit: {max_profit:.1f}% per pair"
    ws[f'B{row}'].fill = PatternFill(start_color=profit_range_color, end_color=profit_range_color, fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
    row += 1
    
    # Check 3: Fit Quality
    buy_quality = fit_metrics.get('fit_quality', 'unknown')
    sell_quality = fit_metrics_sell.get('fit_quality', 'unknown')
    
    buy_fit_status = "PASS" if buy_quality in ['excellent', 'good'] else "WARNING" if buy_quality == 'fair' else "FAIL"
    sell_fit_status = "PASS" if sell_quality in ['excellent', 'good'] else "WARNING" if sell_quality == 'fair' else "FAIL"
    
    buy_fit_color = "00FF00" if buy_fit_status == "PASS" else "FFA500" if buy_fit_status == "WARNING" else "FF0000"
    sell_fit_color = "00FF00" if sell_fit_status == "PASS" else "FFA500" if sell_fit_status == "WARNING" else "FF0000"
    
    ws[f'A{row}'] = "3. Buy-Side Fit Quality"
    ws[f'B{row}'] = buy_fit_status
    ws[f'C{row}'] = f"Quality: {buy_quality}"
    ws[f'B{row}'].fill = PatternFill(start_color=buy_fit_color, end_color=buy_fit_color, fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
    row += 1
    
    ws[f'A{row}'] = "4. Sell-Side Fit Quality"
    ws[f'B{row}'] = sell_fit_status
    ws[f'C{row}'] = f"Quality: {sell_quality}"
    ws[f'B{row}'].fill = PatternFill(start_color=sell_fit_color, end_color=sell_fit_color, fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
    row += 1
    
    # Check 5: Position Size Reasonableness
    max_position = paired_orders_df['buy_notional'].max()
    avg_position = paired_orders_df['buy_notional'].mean()
    position_ratio = max_position / avg_position if avg_position > 0 else 0
    
    position_status = "PASS" if position_ratio <= 3 else "WARNING" if position_ratio <= 5 else "FAIL"
    position_color = "00FF00" if position_status == "PASS" else "FFA500" if position_status == "WARNING" else "FF0000"
    
    ws[f'A{row}'] = "5. Position Size Distribution"
    ws[f'B{row}'] = position_status
    ws[f'C{row}'] = f"Max/Avg ratio: {position_ratio:.1f}"
    ws[f'B{row}'].fill = PatternFill(start_color=position_color, end_color=position_color, fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
    row += 1
    
    row += 1
    
    # Recommendations
    ws[f'A{row}'] = "RECOMMENDATIONS"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    recommendations = []
    
    if len(unprofitable_pairs) > 0:
        recommendations.append("• Review unprofitable pairs and adjust profit targets")
    
    if max_profit > 50:
        recommendations.append("• Consider reducing maximum profit targets for realism")
    
    if buy_quality in ['poor', 'fair']:
        recommendations.append("• Improve buy-side data quality or adjust Weibull parameters")
    
    if sell_quality in ['poor', 'fair']:
        recommendations.append("• Improve sell-side data quality or adjust Weibull parameters")
    
    if position_ratio > 3:
        recommendations.append("• Consider more uniform position sizing")
    
    if not recommendations:
        recommendations.append("• All checks passed - strategy appears sound")
    
    for rec in recommendations:
        ws[f'A{row}'] = rec
        row += 1
    
    # Format columns
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40


if __name__ == "__main__":
    # Test with sample data
    from order_builder import export_orders_csv
    
    depths = np.array([0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0])
    allocations = np.array([50, 60, 70, 80, 90, 100, 110, 120, 130, 140])
    theta, p = 2.5, 1.2
    fit_metrics = {'r_squared': 0.98, 'rmse': 0.001}
    budget = 1000.0
    current_price = 100.0
    
    orders_df = pd.DataFrame({
        'rung': range(1, 11),
        'depth_pct': depths,
        'limit_price': current_price * (1 - depths / 100),
        'quantity': allocations / (current_price * (1 - depths / 100)),
        'notional': allocations,
        'allocation': allocations
    })
    
    create_excel_workbook(orders_df, depths, allocations, theta, p, 
                         fit_metrics, budget, current_price)
    export_orders_csv(orders_df)
